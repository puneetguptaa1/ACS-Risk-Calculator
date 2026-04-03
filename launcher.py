#!/usr/bin/env python3
"""
Interactive launcher for the NSQIP Risk Calculator batch script.

Prompts the user for:
  1. Excel file path
  2. Sheet name
  3. Row range (start / end)

Then runs nsqip_batch.py with the correct arguments.
"""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed. Run the setup script (run.sh or run.bat) first.")
    sys.exit(1)


SCRIPT_DIR = Path(__file__).resolve().parent
BATCH_SCRIPT = SCRIPT_DIR / "nsqip_batch.py"


def prompt(msg: str, default: str = "") -> str:
    """Show a prompt with an optional default; return stripped input."""
    if default:
        raw = input(f"{msg} [{default}]: ").strip()
        return raw if raw else default
    while True:
        raw = input(f"{msg}: ").strip()
        if raw:
            return raw
        print("  (cannot be empty)")


def main() -> None:
    print()
    print("=" * 60)
    print("  NSQIP Risk Calculator -- Batch Automation")
    print("=" * 60)
    print()

    # --- Excel file path ---
    print("Drag and drop your Excel file here, or type the path:")
    raw_path = prompt("  Excel file path").strip("'\"")
    excel_path = Path(raw_path).expanduser().resolve()
    if not excel_path.is_file():
        print(f"\nERROR: File not found: {excel_path}")
        sys.exit(1)
    if not excel_path.suffix.lower().endswith((".xlsx", ".xlsm", ".xltx")):
        print(f"\nWARNING: {excel_path.name} may not be a valid Excel file.")

    # --- List sheets ---
    print(f"\nOpening {excel_path.name}...")
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"\nERROR: Could not open workbook: {e}")
        sys.exit(1)

    sheets = wb.sheetnames
    print(f"\nAvailable sheets ({len(sheets)}):")
    for i, name in enumerate(sheets, 1):
        print(f"  {i}. {name}")
    print()

    # --- Sheet name ---
    sheet_input = prompt("Enter sheet name (or number from the list above)")
    if sheet_input.isdigit():
        idx = int(sheet_input) - 1
        if 0 <= idx < len(sheets):
            sheet_name = sheets[idx]
        else:
            print(f"\nERROR: Number {sheet_input} is out of range.")
            sys.exit(1)
    else:
        matched = [s for s in sheets if s.lower() == sheet_input.lower()]
        if matched:
            sheet_name = matched[0]
        else:
            print(f"\nERROR: Sheet '{sheet_input}' not found. Available: {sheets}")
            sys.exit(1)

    ws = wb[sheet_name]
    max_row = ws.max_row
    print(f"\nSheet '{sheet_name}' has {max_row - 1} data rows (rows 2-{max_row}).")
    wb.close()

    # --- Row range ---
    print()
    start_row = int(prompt("Start row", "2"))
    end_row = int(prompt("End row", str(max_row)))

    if start_row < 2:
        print("\nERROR: Start row must be >= 2 (row 1 is headers).")
        sys.exit(1)
    if end_row < start_row:
        print(f"\nERROR: End row ({end_row}) is less than start row ({start_row}).")
        sys.exit(1)

    # --- Confirmation ---
    json_name = f"{excel_path.stem}_{sheet_name.lower()}_risks.json"
    print()
    print("-" * 60)
    print(f"  Excel file : {excel_path.name}")
    print(f"  Sheet      : {sheet_name}")
    print(f"  Rows       : {start_row} - {end_row}")
    print(f"  JSON output: {json_name}")
    print("-" * 60)
    confirm = input("\nProceed? [Y/n]: ").strip().lower()
    if confirm and confirm not in ("y", "yes"):
        print("Cancelled.")
        sys.exit(0)

    # --- Run batch script ---
    python = sys.executable
    json_path = excel_path.parent / json_name
    cmd = [
        python, str(BATCH_SCRIPT),
        "--excel", str(excel_path),
        "--sheet", sheet_name,
        "--start-row", str(start_row),
        "--end-row", str(end_row),
    ]

    print(f"\nStarting batch processing...\n")
    result = subprocess.run(cmd)

    if result.returncode != 0:
        print(f"\nBatch script exited with error (code {result.returncode}).")
        sys.exit(result.returncode)

    # --- Write results back to Excel ---
    if json_path.is_file():
        print()
        print("=" * 60)
        print("  Writing results back to Excel...")
        print("=" * 60)
        print()
        write_cmd = [
            python, str(SCRIPT_DIR / "json_to_excel.py"),
            "--json", str(json_path),
            "--excel", str(excel_path),
            "--sheet", sheet_name,
        ]
        write_result = subprocess.run(write_cmd)
        if write_result.returncode == 0:
            print(f"\nResults written to '{excel_path.name}' sheet '{sheet_name}'.")
        else:
            print(f"\nWARNING: Writing to Excel failed (code {write_result.returncode}).")
            print(f"Results are still saved in: {json_path}")
    else:
        print(f"\nNo JSON output found at {json_path} -- nothing to write to Excel.")

    sys.exit(0)


if __name__ == "__main__":
    main()
