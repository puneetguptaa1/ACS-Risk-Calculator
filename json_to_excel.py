#!/usr/bin/env python3
"""
Write JSON risk values back into the Excel sheet.

Reads blank_tongue_nsqip_2024_risks.json (keyed by CASEID) and writes the
15 scraped risk values into the corresponding row of the Anahita sheet,
matching on the CASEID in column B.

Usage:
    python json_to_excel.py
    python json_to_excel.py --json blank_tongue_nsqip_2024_risks.json --excel blank_tongue_nsqip_2024.xlsx
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# JSON key -> Excel column letter
JSON_TO_COL: dict[str, str] = {
    "serious_complication_no_adjustment":          "X",
    "serious_complication_somewhat_higher":        "Y",
    "serious_complication_significantly_higher":   "Z",
    "any_complication_no_adjustment":              "AA",
    "any_complication_somewhat_higher":            "AB",
    "any_complication_significantly_higher":       "AC",
    "return_to_or_no_adjustment":                  "AD",
    "return_to_or_somewhat_higher":                "AE",
    "return_to_or_significantly_higher":           "AF",
    "surgical_site_infection_no_adjustment":       "AG",
    "surgical_site_infection_somewhat_higher":     "AH",
    "surgical_site_infection_significantly_higher":"AI",
    "pneumonia_no_adjustment":                     "AJ",
    "pneumonia_somewhat_higher":                   "AK",
    "pneumonia_significantly_higher":              "AL",
}

# Pre-compute integer column indices once
JSON_TO_COLIDX: dict[str, int] = {
    k: column_index_from_string(v) for k, v in JSON_TO_COL.items()
}


def normalise_case_id(raw) -> str:
    """Match the same normalisation used by the scraper's get_case_id()."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def main() -> None:
    root = Path(__file__).resolve().parent

    ap = argparse.ArgumentParser(description="Write JSON risk values into the Excel sheet.")
    ap.add_argument("--json",  type=Path, default=root / "blank_tongue_nsqip_2024_risks.json")
    ap.add_argument("--excel", type=Path, default=root / "blank_tongue_nsqip_2024.xlsx")
    ap.add_argument("--sheet", default="Anahita")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )
    log = logging.getLogger("json_to_excel")

    json_path  = args.json.expanduser().resolve()
    excel_path = args.excel.expanduser().resolve()

    if not json_path.is_file():
        log.error("JSON file not found: %s", json_path)
        sys.exit(1)
    if not excel_path.is_file():
        log.error("Excel file not found: %s", excel_path)
        sys.exit(1)

    risks: dict[str, dict] = json.loads(json_path.read_text())
    log.info("Loaded %d entries from %s", len(risks), json_path.name)

    wb = load_workbook(excel_path)

    # Case-insensitive sheet lookup
    ws = None
    for name in wb.sheetnames:
        if name.lower() == args.sheet.lower():
            ws = wb[name]
            break
    if ws is None:
        log.error("Sheet %r not found (available: %s)", args.sheet, wb.sheetnames)
        sys.exit(1)

    written = skipped = missing = 0

    for row in range(2, ws.max_row + 1):
        case_id = normalise_case_id(ws.cell(row=row, column=2).value)
        if not case_id:
            continue

        entry = risks.get(case_id)
        if entry is None:
            log.debug("Row %d  CASEID=%s  not in JSON — skipping", row, case_id)
            missing += 1
            continue

        for json_key, col_idx in JSON_TO_COLIDX.items():
            value = entry.get(json_key)
            ws.cell(row=row, column=col_idx).value = value

        log.info("Row %d  CASEID=%s  written", row, case_id)
        written += 1

    wb.save(excel_path)
    log.info(
        "Done. %d rows written, %d rows skipped (no JSON entry), %d rows had no CASEID.",
        written, missing, skipped,
    )


if __name__ == "__main__":
    main()
