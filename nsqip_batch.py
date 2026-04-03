#!/usr/bin/env python3
"""
Batch fill ACS NSQIP Risk Calculator from Excel and save 15 scraped
"Your risk" values (columns C-W) to a JSON file keyed by CASEID.

Works with any sheet name — the JSON output is automatically named
<excel_stem>_<sheet>_risks.json so each sheet gets its own file.

Setup (from project directory):
  source venv/bin/activate
  pip install -r requirements.txt
  playwright install chrome

Examples:
  python nsqip_batch.py --excel data.xlsx --sheet Anahita --end-row 9999

  python nsqip_batch.py --excel data.xlsx --sheet Charbel --start-row 2 --end-row 50

  python nsqip_batch.py --dry-run --sheet Anahita

If the disclaimer/reCAPTCHA home page appears, the script waits until you complete it and reach
Patient Information. Optional: open the home page first with
  --entry-url https://riskcalculator.facs.org/RiskCalculator/index.jsp
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeout
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

PATIENT_URL = "https://riskcalculator.facs.org/RiskCalculator/PatientInfo.jsp"
INDEX_URL = "https://riskcalculator.facs.org/RiskCalculator/index.jsp"

_USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"


def ensure_patient_info_page(page, log: logging.Logger, entry_url: str) -> None:
    """
    Navigate to entry URL. If the disclaimer/captcha home page is active (#RkCalHomeForm),
    wait for the user to complete reCAPTCHA, check the disclaimer, and click Continue;
    then wait until the patient form (#RkCalForm) is shown.
    """
    page.goto(entry_url, wait_until="domcontentloaded", timeout=120_000)
    try:
        page.wait_for_selector("#RkCalForm, #RkCalHomeForm", timeout=120_000)
    except PlaywrightTimeout as e:
        raise RuntimeError(
            "Page did not load patient form or disclaimer/captcha home "
            "(expected #RkCalForm or #RkCalHomeForm)."
        ) from e

    if page.locator("#RkCalForm").is_visible():
        log.debug("Patient Information page is active (#RkCalForm).")
        return

    if page.locator("#RkCalHomeForm").is_visible():
        log.info(
            "Disclaimer / reCAPTCHA page is active. In the browser: (1) complete the captcha, "
            "(2) check \u201cI have read the disclaimer and risk calculator permitted use statements below.\u201d, "
            "(3) click Continue. Waiting (up to 15 minutes) for the patient information form\u2026"
        )
        page.wait_for_selector("#RkCalForm", state="visible", timeout=900_000)
        log.info("Patient Information form is ready; continuing.")
        return

    raise RuntimeError(
        "Unrecognized calculator page: neither #RkCalForm nor #RkCalHomeForm is visible."
    )


def handle_mid_session_captcha(page, log: logging.Logger) -> bool:
    """
    Call after any navigation that might land on the mid-session threshold captcha page.
    Returns True if a captcha was detected and waited through; False if already on the right page.
    """
    if page.locator("#RkCalHomeForm").is_visible():
        if "Threshold" in page.title():
            msg = (
                "Mid-session captcha (Threshold page) appeared after submitting the form. "
                "Complete the reCAPTCHA in the browser and click Continue. "
                "Waiting (up to 15 minutes)..."
            )
        else:
            msg = (
                "Captcha page appeared after submitting the form. "
                "Complete it in the browser and click Continue. "
                "Waiting (up to 15 minutes)..."
            )
        log.info(msg)
        page.wait_for_selector("#RkCalForm, #ChartDIVChart", state="visible", timeout=900_000)
        log.info("Captcha resolved; continuing.")
        return True
    return False

INPUT_COL_START = 3
INPUT_COL_END = 23

OUTPUT_COLS: Dict[str, Tuple[str, str, str]] = {
    "serious": ("X", "Y", "Z"),
    "any": ("AA", "AB", "AC"),
    "return_or": ("AD", "AE", "AF"),
    "ssi": ("AG", "AH", "AI"),
    "pneumonia": ("AJ", "AK", "AL"),
}

OUTCOME_SCRAPE_ORDER = ("serious", "any", "pneumonia", "ssi", "return_or")


def resolve_sheet(workbook, name: str):
    for sn in workbook.sheetnames:
        if sn.lower() == name.strip().lower():
            return workbook[sn]
    raise KeyError(f"No sheet named {name!r} (have {workbook.sheetnames})")


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


CPT_SUBSTITUTIONS: Dict[str, str] = {
    "20969": "20962",
}


def cpt_to_int_str(val: Any) -> str:
    if val is None or (isinstance(val, str) and not val.strip()):
        return ""
    if isinstance(val, (int, float)):
        code = str(int(round(val)))
    else:
        s = str(val).strip()
        if not s:
            return ""
        code = str(int(round(float(s.replace(",", "")))))
    return CPT_SUBSTITUTIONS.get(code, code)


def yn_binary(val: Any, yes_value: str = "1", no_value: str = "0") -> str:
    t = _norm(val).lower()
    if not t:
        return no_value
    if t in ("y", "yes", "true", "1", "si"):
        return yes_value
    if t in ("n", "no", "false", "0"):
        return no_value
    if "yes" in t or t.startswith("y"):
        return yes_value
    return no_value


def gender_value(val: Any) -> str:
    t = _norm(val).lower()
    if t in ("1", "m", "male", "man"):
        return "1"
    return "0"


def functional_status_value(val: Any) -> str:
    t = _norm(val).lower()
    if "total" in t:
        return "3"
    if "partial" in t:
        return "2"
    return "1"


def emergency_value(val: Any) -> str:
    return yn_binary(val, "1", "0")


def asa_class_value(val: Any) -> str:
    t = _norm(val)
    m = re.match(r"^\s*(\d)", t)
    if m:
        d = int(m.group(1))
        if 1 <= d <= 5:
            return str(d)
    tl = t.lower()
    if "healthy" in tl or "class 1" in tl:
        return "1"
    if "mild" in tl:
        return "2"
    if "severe" in tl and "life" not in tl and "moribund" not in tl:
        return "3"
    if "life" in tl or "constant threat" in tl:
        return "4"
    if "moribund" in tl:
        return "5"
    return "1"


def systemic_sepsis_value(val: Any) -> str:
    t = _norm(val).lower()
    if not t or t in ("n", "no", "none"):
        return "1"
    if "sirs" in t:
        return "2"
    if "shock" in t:
        return "4"
    if "sepsis" in t:
        return "3"
    return "1"


def diabetes_value(val: Any) -> str:
    t = _norm(val).lower()
    if not t or t in ("n", "no", "none"):
        return "1"
    elif "oral" in t or "non-insulin" in t or "non insulin" in t:
        return "2"
    elif "insulin" in t:
        return "3"
    elif yn_binary(val) == "1":
        return "2"
    return "1"


def history_copd_value(val: Any) -> str:
    return yn_binary(val, "1", "0")


def parse_risk_text(raw: str) -> Optional[float]:
    if not raw:
        return None
    s = raw.strip().replace("%", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        m = re.search(r"[\d.]+\s*%?", raw)
        if m:
            try:
                return float(m.group(0).replace("%", ""))
            except ValueError:
                pass
    return None


def norm_title(t: str) -> str:
    return " ".join(t.replace("\n", " ").split()).lower()


def classify_outcome(title: str) -> Optional[str]:
    t = norm_title(title)
    if "any complication" in t:
        return "any"
    if "serious complication" in t:
        return "serious"
    if "surgical site infection" in t or ("site infection" in t and "surgical" in t):
        return "ssi"
    if "pneumonia" in t:
        return "pneumonia"
    if "return" in t and ("operating" in t or re.search(r"\bto the or\b|\bto or\b", t)):
        return "return_or"
    return None


def scrape_main_chart_your_risk(page) -> Dict[str, Optional[float]]:
    page.wait_for_selector("#ChartDIVChart svg.bullet", timeout=120_000)

    results = page.evaluate("""() => {
        const out = [];
        const svgs = document.querySelectorAll('#ChartDIVChart svg.bullet');
        for (const svg of svgs) {
            const titleEl = svg.querySelector('text.title');
            const riskEl  = svg.querySelector('text.titleEstRisk');
            out.push({
                title: titleEl ? titleEl.textContent : '',
                risk:  riskEl  ? riskEl.textContent  : ''
            });
        }
        return out;
    }""")

    found: Dict[str, Optional[float]] = {k: None for k in OUTCOME_SCRAPE_ORDER}
    for item in results:
        key = classify_outcome(item.get("title", ""))
        if key and found[key] is None:
            found[key] = parse_risk_text(item.get("risk", ""))
    return found


def collect_all_adjustment_levels(page, log) -> Dict[str, List[Optional[float]]]:
    """For surgeon levels 1,2,3 return outcome_key -> [v1,v2,v3]."""
    result: Dict[str, List[Optional[float]]] = {k: [None, None, None] for k in OUTCOME_SCRAPE_ORDER}
    page.wait_for_selector("#SurgeonAdjustmentRisk", timeout=60_000)
    page.wait_for_selector("#ChartDIVChart svg.bullet", timeout=120_000)
    page.wait_for_timeout(150)

    for idx, level in enumerate(["1", "2", "3"]):
        if idx > 0:
            try:
                with page.expect_response(
                    lambda r: r.request.method == "POST"
                    and "RiskCalcServlet" in r.url
                    and (r.request.post_data or "").find("action=getRiskCalcResults") >= 0,
                    timeout=120_000,
                ):
                    page.select_option("#SurgeonAdjustmentRisk", level)
            except PlaywrightTimeout:
                log.warning("Timeout waiting for getRiskCalcResults after adjustment %s", level)
        page.wait_for_timeout(150)
        chunk = scrape_main_chart_your_risk(page)
        missing = [k for k in OUTCOME_SCRAPE_ORDER if chunk.get(k) is None]
        if missing:
            log.warning("Missing outcomes after level %s: %s. Titles dump may be needed.", level, missing)
        for k in OUTCOME_SCRAPE_ORDER:
            result[k][idx] = chunk.get(k)
    return result


def get_case_id(ws, row: int) -> str:
    """Return CASEID from column B, stripped of trailing '.0' if numeric string."""
    raw = ws.cell(row=row, column=2).value
    if raw is None:
        return f"row_{row}"
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def row_tuple_to_form(ws, row: int) -> Dict[str, Any]:
    """Read Excel row -> dict of form field IDs."""
    vals = [ws.cell(row=row, column=c).value for c in range(INPUT_COL_START, INPUT_COL_END + 1)]
    (
        cpt, age, sex, func, emerg, asa, steroid, ascites, sepsis, vent,
        cancer, diabetes, htn, chf, o2, smoke, copd, dialysis, aki, height, weight,
    ) = (vals + [None] * 21)[:21]

    return {
        "CPT": cpt,
        "AgeGroup": age,
        "Gender": sex,
        "FunctionalStatus": func,
        "EmergencyCase": emerg,
        "ASAClass": asa,
        "SteroidUse": steroid,
        "Ascites": ascites,
        "SystemicSepsis": sepsis,
        "VentilatorDependent": vent,
        "DisseminatedCancer": cancer,
        "Diabetes": diabetes,
        "Hypertension": htn,
        "CongestiveHeartFailure": chf,
        "OxygenSupport": o2,
        "CurrentSmoker": smoke,
        "HistoryCOPD": copd,
        "Dialysis": dialysis,
        "AcuteRenalFailure": aki,
        "PatientHeight": height,
        "PatientWeight": weight,
    }


def _fv(label: str, raw: Any, mapped: str, log) -> str:
    """Log a field value (raw Excel -> mapped calculator value) and return the mapped value."""
    log.info("  %-28s  raw=%-25s -> %s", label, repr(raw), mapped)
    return mapped


def fill_patient_form(page, data: Dict[str, Any], log):
    page.click("#btnReset")
    page.wait_for_timeout(200)

    log.info("--- Filling form fields ---")

    page.uncheck("#Surgical")
    page.uncheck("#NonOperative")
    page.uncheck("#NoOtherOptions")
    log.info("  %-28s  (left blank -- no Excel data)", "TreatmentOptions")

    age_raw = data.get("AgeGroup")
    age_str = str(int(float(str(age_raw).replace(",", "")))) if age_raw is not None and str(age_raw).strip() else ""
    log.info("  %-28s  raw=%-25s -> %s", "AgeGroup", repr(age_raw), age_str)
    if age_str:
        page.fill("#AgeGroup", age_str)

    gender_raw = data.get("Gender")
    gender_v = gender_value(gender_raw)
    page.select_option("#Gender", _fv("Gender", gender_raw, gender_v, log))

    func_raw = data.get("FunctionalStatus")
    func_v = functional_status_value(func_raw)
    page.select_option("#FunctionalStatus", _fv("FunctionalStatus", func_raw, func_v, log))

    emerg_raw = data.get("EmergencyCase")
    emerg_v = emergency_value(emerg_raw)
    page.select_option("#EmergencyCase", _fv("EmergencyCase", emerg_raw, emerg_v, log))

    asa_raw = data.get("ASAClass")
    asa_v = asa_class_value(asa_raw)
    page.select_option("#ASAClass", _fv("ASAClass", asa_raw, asa_v, log))

    steroid_raw = data.get("SteroidUse")
    steroid_v = yn_binary(steroid_raw, "1", "0")
    page.select_option("#SteroidUse", _fv("SteroidUse", steroid_raw, steroid_v, log))

    ascites_raw = data.get("Ascites")
    ascites_v = yn_binary(ascites_raw, "1", "0")
    page.select_option("#Ascites", _fv("Ascites", ascites_raw, ascites_v, log))

    sepsis_raw = data.get("SystemicSepsis")
    sepsis_v = systemic_sepsis_value(sepsis_raw)
    page.select_option("#SystemicSepsis", _fv("SystemicSepsis", sepsis_raw, sepsis_v, log))

    vent_raw = data.get("VentilatorDependent")
    vent_v = yn_binary(vent_raw, "1", "0")
    page.select_option("#VentilatorDependent", _fv("VentilatorDependent", vent_raw, vent_v, log))

    cancer_raw = data.get("DisseminatedCancer")
    cancer_v = yn_binary(cancer_raw, "1", "0")
    page.select_option("#DisseminatedCancer", _fv("DisseminatedCancer", cancer_raw, cancer_v, log))

    diabetes_raw = data.get("Diabetes")
    diabetes_v = diabetes_value(diabetes_raw)
    page.select_option("#Diabetes", _fv("Diabetes", diabetes_raw, diabetes_v, log))

    htn_raw = data.get("Hypertension")
    htn_v = yn_binary(htn_raw, "1", "0")
    page.select_option("#Hypertension", _fv("Hypertension", htn_raw, htn_v, log))

    chf_raw = data.get("CongestiveHeartFailure")
    chf_v = yn_binary(chf_raw, "1", "0")
    page.select_option("#CongestiveHeartFailure", _fv("CongestiveHeartFailure", chf_raw, chf_v, log))

    o2_raw = data.get("OxygenSupport")
    o2_v = yn_binary(o2_raw, "1", "0")
    page.select_option("#OxygenSupport", _fv("OxygenSupport", o2_raw, o2_v, log))

    smoke_raw = data.get("CurrentSmoker")
    smoke_v = yn_binary(smoke_raw, "1", "0")
    page.select_option("#CurrentSmoker", _fv("CurrentSmoker", smoke_raw, smoke_v, log))

    copd_raw = data.get("HistoryCOPD")
    copd_v = history_copd_value(copd_raw)
    page.select_option("#HistoryCOPD", _fv("HistoryCOPD", copd_raw, copd_v, log))

    dialysis_raw = data.get("Dialysis")
    dialysis_v = yn_binary(dialysis_raw, "1", "0")
    page.select_option("#Dialysis", _fv("Dialysis", dialysis_raw, dialysis_v, log))

    aki_raw = data.get("AcuteRenalFailure")
    aki_v = yn_binary(aki_raw, "1", "0")
    page.select_option("#AcuteRenalFailure", _fv("AcuteRenalFailure", aki_raw, aki_v, log))

    h = data.get("PatientHeight")
    w = data.get("PatientWeight")
    h_valid = h is not None and str(h).strip() and float(str(h)) != -99
    w_valid = w is not None and str(w).strip() and float(str(w)) != -99
    h_fill = str(int(float(str(h).replace(",", "")))) if h_valid and w_valid else ""
    w_fill = str(int(float(str(w).replace(",", "")))) if w_valid and h_valid else ""
    log.info("  %-28s  raw=%-25s -> %s", "PatientHeight (in)", repr(h), h_fill or "(blank)")
    log.info("  %-28s  raw=%-25s -> %s", "PatientWeight (lb)", repr(w), w_fill or "(blank)")
    page.fill("#PatientHeight", h_fill)
    page.fill("#PatientWeight", w_fill)

    cpt_s = cpt_to_int_str(data.get("CPT"))
    if not cpt_s:
        raise ValueError("Empty CPT")
    log.info("  %-28s  raw=%-25s -> %s", "CPT", repr(data.get("CPT")), cpt_s)

    proc = page.locator("#txtProcedure")
    proc.click()
    proc.type(cpt_s, delay=50)
    proc.press("Enter")
    page.wait_for_timeout(2_000)

    cpt_hidden = page.input_value("#CPT").strip()
    log.info("  %-28s  confirmed #CPT=%s", "CPT (hidden field)", cpt_hidden)
    if not cpt_hidden:
        raise RuntimeError(f"CPT not selected after autocomplete (typed {cpt_s})")
    log.info("--- Form fields done ---")

    if page.is_visible("#ipBannedMsg") and page.locator("#ipBannedMsg").inner_text().strip():
        raise RuntimeError(page.locator("#ipBannedMsg").inner_text())


def collected_to_json_entry(collected: Dict[str, List[Optional[float]]]) -> Dict[str, Any]:
    """Convert collected risks to a flat JSON-friendly dict."""
    level_keys = ("no_adjustment", "somewhat_higher", "significantly_higher")
    outcome_names = {
        "serious":   "serious_complication",
        "any":       "any_complication",
        "pneumonia": "pneumonia",
        "ssi":       "surgical_site_infection",
        "return_or": "return_to_or",
    }
    entry: Dict[str, Any] = {}
    for key, name in outcome_names.items():
        vals = collected.get(key, [None, None, None])
        for i, level in enumerate(level_keys):
            entry[f"{name}_{level}"] = vals[i] if i < len(vals) else None
    return entry


def print_dry_run_results(
    row: int,
    collected: Dict[str, List[Optional[float]]],
    data: Dict[str, Any],
) -> None:
    """Print scraped risks to stdout (--dry-run)."""
    cpt_s = cpt_to_int_str(data.get("CPT"))
    labels = ("1_no_adjustment", "2_somewhat_higher", "3_significantly_higher")
    titles = {
        "serious": "Serious complication",
        "any": "Any complication",
        "pneumonia": "Pneumonia",
        "ssi": "Surgical site infection",
        "return_or": "Return to OR",
    }
    print(f"\n=== Dry run: Excel row {row} (CPT {cpt_s}) ===\n")
    for key in OUTCOME_SCRAPE_ORDER:
        vals = collected.get(key, [None, None, None])
        print(titles.get(key, key))
        for i, lab in enumerate(labels):
            v = vals[i] if i < len(vals) else None
            print(f"  {lab}: {v}")
        print()
    print("=== End dry run (workbook not saved) ===\n")


def find_first_row_with_cpt(ws, start: int, end: int) -> Optional[int]:
    for row in range(start, end + 1):
        data = row_tuple_to_form(ws, row)
        if cpt_to_int_str(data.get("CPT")):
            return row
    return None


def process_row(
    page,
    ws,
    row: int,
    log,
    pause_captcha: bool,
    captcha_done: List[bool],
    entry_url: str = PATIENT_URL,
) -> Optional[Dict[str, List[Optional[float]]]]:
    data = row_tuple_to_form(ws, row)
    cpt_s = cpt_to_int_str(data["CPT"])
    if not cpt_s:
        log.info("Row %s: skip (empty CPT)", row)
        return None

    ensure_patient_info_page(page, log, entry_url)

    if pause_captcha and not captcha_done[0]:
        log.info("Optional pause (--wait-captcha): press Enter in this terminal when ready to fill the form.")
        input("[Enter] ")
        captcha_done[0] = True

    def on_dialog(diag):
        msg = diag.message
        diag.accept()
        raise RuntimeError(f"Browser alert: {msg}")

    page.once("dialog", on_dialog)

    fill_patient_form(page, data, log)

    with page.expect_navigation(wait_until="domcontentloaded", timeout=180_000):
        page.click("#btnNext")

    if handle_mid_session_captcha(page, log):
        page.wait_for_selector("#ChartDIVChart, #RkCalForm, #ipBannedMsg", timeout=180_000)

    page.wait_for_selector("#ChartDIVChart, #ipBannedMsg, div.content-wrapper", timeout=180_000)
    if page.is_visible("#ipBannedMsg"):
        txt = page.locator("#ipBannedMsg").inner_text().strip()
        if txt:
            raise RuntimeError(f"IP/limit message: {txt}")

    page.wait_for_selector("#SurgeonAdjustmentRisk", timeout=120_000)
    collected = collect_all_adjustment_levels(page, log)
    log.info("Row %s: scraped risks CPT=%s %s", row, cpt_s, collected)
    return collected


def main():
    ap = argparse.ArgumentParser(
        description="Fill NSQIP Risk Calculator from Excel and save scraped risks to JSON keyed by CASEID.",
        epilog="First-time setup: pip install -r requirements.txt && playwright install chrome",
    )
    root = Path(__file__).resolve().parent
    ap.add_argument(
        "--excel",
        type=Path,
        default=root / "blank_tongue_nsqip_2024.xlsx",
        help="Path to workbook",
    )
    ap.add_argument("--sheet", required=True, help="Sheet name (case-insensitive)")
    ap.add_argument("--start-row", type=int, default=2)
    ap.add_argument(
        "--end-row",
        type=int,
        default=None,
        help="Inclusive last row to process. Default: start-row + 4 (5 rows). Set higher for full run.",
    )
    ap.add_argument(
        "--json-output",
        type=Path,
        default=None,
        help="Path for JSON results file (default: <excel_stem>_<sheet>_risks.json)",
    )
    ap.add_argument("--headless", action="store_true", help="Headless browser (may block captcha)")
    ap.add_argument(
        "--profile-dir",
        type=Path,
        default=None,
        help=(
            "Path to a persistent browser profile directory (created on first run). "
            "Reusing the same directory preserves cookies/session between runs so "
            "you solve the captcha far less often. "
            "Default: <project>/browser_profile"
        ),
    )
    ap.add_argument(
        "--wait-captcha",
        action="store_true",
        help="After reaching PatientInfo, pause once for Enter (optional; captcha pages handled automatically)",
    )
    ap.add_argument(
        "--entry-url",
        default=PATIENT_URL,
        help=(
            f"First page to open (default: {PATIENT_URL}). "
            f"Use {INDEX_URL} to start on the disclaimer/captcha home page."
        ),
    )
    ap.add_argument("--delay-seconds", type=float, default=2.0, help="Pause between rows (seconds)")
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Process first row with CPT, print JSON to stdout, do not write file",
    )
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )
    log = logging.getLogger("nsqip")

    excel_path = args.excel.expanduser().resolve()
    if not excel_path.is_file():
        log.error("Workbook not found: %s", excel_path)
        sys.exit(1)

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = resolve_sheet(wb, args.sheet)

    end_row = args.end_row
    if end_row is None:
        end_row = args.start_row + 4
    end_row = min(end_row, ws.max_row)
    log.info("Processing rows %s - %s (of %s max)", args.start_row, end_row, ws.max_row)

    json_path = args.json_output
    if json_path is None:
        json_path = excel_path.parent / f"{excel_path.stem}_{args.sheet.lower()}_risks.json"
    json_path = Path(json_path).expanduser().resolve()

    captcha_flag = [False]

    results: Dict[str, Any] = {}
    if json_path.is_file():
        try:
            results = json.loads(json_path.read_text())
            log.info("Loaded %d existing entries from %s", len(results), json_path)
        except json.JSONDecodeError:
            log.warning("Could not parse existing %s; starting fresh", json_path)

    profile_dir = args.profile_dir
    if profile_dir is None:
        profile_dir = excel_path.parent / "browser_profile"
    profile_dir = Path(profile_dir).expanduser().resolve()
    profile_dir.mkdir(parents=True, exist_ok=True)
    log.info("Browser profile: %s", profile_dir)

    _browser_kwargs = dict(
        channel="chrome",
        headless=args.headless,
        viewport={"width": 1280, "height": 1024},
        locale="en-US",
        extra_http_headers={
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        },
        args=["--disable-blink-features=AutomationControlled"],
    )

    if args.dry_run:
        run_row = find_first_row_with_cpt(ws, args.start_row, end_row)
        if run_row is None:
            log.error("No row with CPT between %s and %s", args.start_row, end_row)
            sys.exit(1)
        case_id = get_case_id(ws, run_row)
        data = row_tuple_to_form(ws, run_row)
        log.info("Dry run: Excel row %s  CASEID=%s", run_row, case_id)
        with Stealth().use_sync(sync_playwright()) as p:
            context = p.chromium.launch_persistent_context(
                user_data_dir=str(profile_dir),
                **_browser_kwargs,
            )
            page = context.new_page()
            try:
                collected = process_row(
                    page, ws, run_row, log,
                    pause_captcha=args.wait_captcha,
                    captcha_done=captcha_flag,
                    entry_url=args.entry_url,
                )
            finally:
                context.close()
        if collected:
            print_dry_run_results(run_row, collected, data)
            entry = collected_to_json_entry(collected)
            entry["cpt"] = cpt_to_int_str(data.get("CPT"))
            print("\nJSON entry:")
            print(json.dumps({case_id: entry}, indent=2))
        else:
            log.error("Dry run produced no results")
            sys.exit(1)
        return

    with Stealth().use_sync(sync_playwright()) as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(profile_dir),
            **_browser_kwargs,
        )
        page = context.new_page()

        try:
            for row in range(args.start_row, end_row + 1):
                case_id = get_case_id(ws, row)
                try:
                    collected = process_row(
                        page, ws, row, log,
                        pause_captcha=args.wait_captcha,
                        captcha_done=captcha_flag,
                        entry_url=args.entry_url,
                    )
                    if collected is not None:
                        entry = collected_to_json_entry(collected)
                        entry["cpt"] = cpt_to_int_str(row_tuple_to_form(ws, row).get("CPT"))
                        results[case_id] = entry
                        json_path.write_text(json.dumps(results, indent=2))
                        log.info("Saved CASEID=%s to %s", case_id, json_path)
                        if args.delay_seconds > 0:
                            time.sleep(args.delay_seconds)
                except Exception as e:
                    log.exception("Row %s (CASEID=%s) failed: %s", row, case_id, e)
        finally:
            context.close()

    json_path.write_text(json.dumps(results, indent=2))
    log.info("Final JSON saved to %s  (%d entries)", json_path, len(results))


if __name__ == "__main__":
    main()
