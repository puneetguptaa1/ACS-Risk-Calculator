#!/usr/bin/env python3
"""
Fill one row from the Excel sheet into the NSQIP calculator, then stop.
The browser stays open so you can review the filled form and click Continue yourself.

Defaults to the Anahita sheet. Use --sheet Charbel for Charbel's rows.

Usage:
    python fill_row.py --row 5
    python fill_row.py --caseid 15397330
    python fill_row.py --row 12 --sheet Charbel
    python fill_row.py --row 12 --entry-url https://riskcalculator.facs.org/RiskCalculator/index.jsp
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

from nsqip_batch import (
    PATIENT_URL,
    INDEX_URL,
    _USER_AGENT,
    resolve_sheet,
    get_case_id,
    row_tuple_to_form,
    cpt_to_int_str,
    fill_patient_form,
    ensure_patient_info_page,
)


def find_row_by_caseid(ws, caseid: str) -> int | None:
    """Return the first row whose column-B CASEID matches (case-insensitive string match)."""
    target = caseid.strip()
    for row in range(2, ws.max_row + 1):
        if get_case_id(ws, row) == target:
            return row
    return None


def main() -> None:
    root = Path(__file__).resolve().parent

    ap = argparse.ArgumentParser(
        description="Fill one Excel row into the NSQIP calculator and leave the browser open."
    )
    ap.add_argument(
        "--excel",
        type=Path,
        default=root / "blank_tongue_nsqip_2024.xlsx",
        help="Path to workbook (default: blank_tongue_nsqip_2024.xlsx)",
    )
    ap.add_argument("--sheet", default="Charbel", help="Sheet name (case-insensitive)")

    target = ap.add_mutually_exclusive_group(required=True)
    target.add_argument("--row",    type=int, help="Excel row number (2 = first data row)")
    target.add_argument("--caseid", type=str, help="CASEID value from column B")

    ap.add_argument(
        "--entry-url",
        default=PATIENT_URL,
        help=f"Page to open first (default: {PATIENT_URL}). "
             f"Use {INDEX_URL} to start on the disclaimer/captcha home page.",
    )
    ap.add_argument("--profile-dir", type=Path, default=None,
                    help="Persistent browser profile directory (default: browser_profile/)")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )
    log = logging.getLogger("fill_row")

    excel_path = args.excel.expanduser().resolve()
    if not excel_path.is_file():
        log.error("Workbook not found: %s", excel_path)
        sys.exit(1)

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = resolve_sheet(wb, args.sheet)

    # Resolve which row to fill
    if args.row is not None:
        row = args.row
        if row < 2 or row > ws.max_row:
            log.error("Row %d is out of range (sheet has rows 2–%d)", row, ws.max_row)
            sys.exit(1)
    else:
        row = find_row_by_caseid(ws, args.caseid)
        if row is None:
            log.error("CASEID %r not found in column B", args.caseid)
            sys.exit(1)

    case_id = get_case_id(ws, row)
    data    = row_tuple_to_form(ws, row)
    cpt_s   = cpt_to_int_str(data.get("CPT"))

    log.info("Excel row %d  |  CASEID=%s  |  CPT=%s", row, case_id, cpt_s or "(empty)")

    if not cpt_s:
        log.error("Row %d has no CPT value — nothing to fill.", row)
        sys.exit(1)

    profile_dir = args.profile_dir or (root / "browser_profile")
    profile_dir = Path(profile_dir).expanduser().resolve()
    profile_dir.mkdir(parents=True, exist_ok=True)

    browser_kwargs = dict(
        channel="chrome",
        headless=False,
        viewport={"width": 1280, "height": 1024},
        locale="en-US",
        extra_http_headers={
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        },
        args=["--disable-blink-features=AutomationControlled"],
    )

    with Stealth().use_sync(sync_playwright()) as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(profile_dir),
            **browser_kwargs,
        )
        page = context.new_page()

        try:
            ensure_patient_info_page(page, log, args.entry_url)
            fill_patient_form(page, data, log)
            log.info("")
            log.info("Form filled for CASEID=%s (row %d, CPT=%s).", case_id, row, cpt_s)
            log.info("The browser is open — review the form and click Continue yourself.")
            log.info("Press Enter here to close the browser when you are done.")
            input()
        except Exception as e:
            log.error("Error: %s", e)
            log.info("Browser left open. Press Enter to close.")
            input()
        finally:
            context.close()


if __name__ == "__main__":
    main()
